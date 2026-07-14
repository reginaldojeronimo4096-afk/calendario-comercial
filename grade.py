# -*- coding: utf-8 -*-
"""
Grade de Ativação: lê o Excel (.xlsm) da grade e mostra as promoções.

- Só considera as abas cujo nome começa com 'LISTA_' e que estão VISÍVEIS
  (abas ocultas são ignoradas de propósito).
- De cada aba pega os dados da promoção (período, link, %) e os produtos.
- Ao subir uma grade nova, SUBSTITUI as listas que vierem no arquivo; as que não
  vierem ficam como histórico (não são apagadas). Ver db.py.
"""
import io
import re

import openpyxl
import pandas as pd
import streamlit as st

import db

# Colunas dos produtos, por posição (1-based) na aba LISTA_:
#   H=SKU, I=DESCRIÇÃO, J=CATEGORIA PLANEJAMENTO, M=LINHA, N=KVIs, T=MECÂNICA,
#   U=SELO, Y=DE, Z=POR, AA=DESCONTO, AF=CICLO PROMO, AG=TIPO
_COLS = {
    "sku": 8, "descricao": 9, "categoria": 10, "linha": 13, "kvis": 14,
    "mecanica": 20, "selo": 21, "preco_de": 25, "preco_por": 26,
    "desconto": 27, "ciclo_promo": 32, "tipo": 33,
}

# Como mostrar cada coluna na tela (ordem e rótulo).
_ROTULOS = {
    "sku": "SKU", "descricao": "Descrição", "categoria": "Categoria",
    "linha": "Linha", "kvis": "KVIs", "mecanica": "Mecânica", "selo": "Selo",
    "preco_de": "DE", "preco_por": "POR", "desconto": "Desconto",
    "ciclo_promo": "Ciclo Promo", "tipo": "Tipo",
}


# ---------------------------------------------------------------------------
# Leitura da planilha
# ---------------------------------------------------------------------------
def _tipo_amigavel(nome: str) -> str:
    """Nome COMPLETO da aba (mantém o prefixo 'LISTA_xx', a pedido), só trocando
    '_' por espaço p/ ficar mais legível.
    Ex.: 'LISTA_20.5_MEIO_AMBIENTE' -> 'LISTA 20.5 MEIO AMBIENTE'.
    (Antes ele APAGAVA o 'LISTA_xx' e mostrava só 'Meio Ambiente'.)"""
    return nome.replace("_", " ").strip() if nome else nome


def _ciclo_periodo(nome_arq: str):
    """Extrai o ciclo (CL_09_2026) e o período (26.05 a 31.05) do nome do arquivo."""
    per = re.search(r"\(([\d.]+ a [\d.]+)\)", nome_arq or "")
    cic = re.search(r"(CL[_ ]?\d+[_ ]?\d{4})", nome_arq or "")
    return (cic.group(1) if cic else "", per.group(1) if per else "")


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:   # NaN
        return ""
    return str(v).strip()


def _fmt_data(v) -> str:
    """Data do Excel -> 'dd/mm/aaaa'. Aceita datetime ou texto (ou vazio)."""
    if v is None or v == "":
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return _txt(v)


def _periodo_acao(b1, c1) -> str:
    """Período da ação a partir de B1/C1 da aba. Se B1 tiver 'CICLO', a ação vale o
    ciclo inteiro. Senão, 'dd/mm/aaaa a dd/mm/aaaa' (ou só uma data, se faltar a outra)."""
    if "CICLO" in _txt(b1).upper():
        return "Todo o ciclo"
    ini, fim = _fmt_data(b1), _fmt_data(c1)
    if ini and fim:
        return f"{ini} a {fim}"
    return ini or fim or ""


def _ciclo_curto(c: str) -> str:
    """Tira o 'CL' do ciclo p/ exibir mais limpo: 'CL_09_2026' -> '09_2026'."""
    return re.sub(r"^CL[_ ]?", "", (c or "").strip())


def ler_grade(file_bytes: bytes, nome_arquivo: str) -> list:
    """Lê o arquivo e devolve [(meta, produtos), ...] das abas LISTA_ visíveis."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ciclo, periodo = _ciclo_periodo(nome_arquivo)
    resultado = []
    for nome in wb.sheetnames:
        if not nome.startswith("LISTA_"):
            continue
        ws = wb[nome]
        if ws.sheet_state != "visible":     # ignora abas ocultas
            continue
        linhas = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=42))

        def celula(lin, col):               # col 1-based
            if len(linhas) > lin and len(linhas[lin]) >= col:
                return linhas[lin][col - 1].value
            return None

        meta = {
            "lista_nome": nome,
            "tipo": _tipo_amigavel(nome),
            "ciclo": ciclo,
            "periodo": periodo,
            "link_lp": _txt(celula(1, 2)),        # B2 = LP
            "periodo_acao": _periodo_acao(celula(0, 2), celula(0, 3)),  # B1 (ou "CICLO") + C1
            "comissao": _num(celula(2, 40)),      # AN3
            "cupom": _num(celula(2, 41)),         # AO3
            "depor": _num(celula(2, 42)),         # AP3
        }
        produtos = []
        for row in linhas[4:]:                    # produtos a partir da linha 5
            if not row[0].value:                  # coluna A (INSERIR SKU) vazia -> ignora
                continue
            p = {"lista_nome": nome}
            for chave, col in _COLS.items():
                v = row[col - 1].value if len(row) >= col else None
                p[chave] = _num(v) if chave in ("preco_de", "preco_por", "desconto") else _txt(v)
            produtos.append(p)
        meta["total_skus"] = len(produtos)
        resultado.append((meta, produtos))
    wb.close()
    return resultado


def salvar_grade(file_bytes: bytes, nome_arquivo: str, marca: str = "natura") -> list:
    """Lê o arquivo e grava no Supabase, na empresa 'marca'. Retorna [(tipo, n), ...]."""
    resumo = []
    for meta, produtos in ler_grade(file_bytes, nome_arquivo):
        db.grade_upsert_lista(meta, marca)
        db.grade_substituir_produtos(meta["lista_nome"], produtos, marca)
        resumo.append((meta["tipo"], meta["total_skus"]))
    return resumo


# ---------------------------------------------------------------------------
# Tela de promoções (página inteira)
# ---------------------------------------------------------------------------
def _pct(v):
    return f"{round(v * 100)}%" if isinstance(v, (int, float)) else "—"


def _excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Produtos")
    return buffer.getvalue()


def _painel_apagar_promocoes(marca: str = "natura") -> None:
    """Expander (admin) p/ APAGAR promoções DESTA empresa: por seleção OU todas."""
    # Cabeçalho do expander em VERMELHO (zona de perigo), sem afetar o "Atualizar
    # Grade": envolvo num container com chave 'exp_apagar' p/ mirar só neste expander.
    st.markdown(
        "<style>"
        ".st-key-exp_apagar [data-testid='stExpander'] summary{"
        "background:#FDECEC !important;color:#C0392B !important;"
        "border:1px solid #E8463A !important;font-weight:700 !important;}"
        ".st-key-exp_apagar [data-testid='stExpander'] summary:hover{"
        "background:#FADBD8 !important;}"
        "</style>",
        unsafe_allow_html=True,
    )
    with st.container(key="exp_apagar"):
        with st.expander("🗑️ Apagar promoções", expanded=False):
            _listas = db.grade_listar_listas(marca)
            if not _listas:
                st.caption("Não há promoções para apagar.")
                return
            # nome amigável (mostrado) -> lista_nome (chave real no banco)
            _map = {_tipo_amigavel(x["lista_nome"]): x["lista_nome"] for x in _listas}
            _opts = sorted(_map.keys())
            # Tira da seleção itens que já não existem (evita erro do multiselect).
            if "apagar_sel" in st.session_state:
                st.session_state.apagar_sel = [
                    s for s in st.session_state.apagar_sel if s in _opts
                ]
            st.caption("Selecione promoções para apagar, ou apague TODAS de uma vez. "
                       "⚠️ Apagar não tem como desfazer.")
            _sel = st.multiselect(
                "Promoções para apagar (uma ou várias):", options=_opts, key="apagar_sel",
            )
            # Botões de apagar em VERMELHO (ação destrutiva). Desabilitado = neutro.
            st.markdown(
                "<style>"
                ".st-key-btn_ap_sel button:not(:disabled),"
                ".st-key-btn_ap_all button:not(:disabled){"
                "background:#FDECEC !important;border:1px solid #E8463A !important;"
                "color:#C0392B !important;font-weight:700 !important;}"
                ".st-key-btn_ap_sel button:not(:disabled):hover,"
                ".st-key-btn_ap_all button:not(:disabled):hover{background:#FADBD8 !important;}"
                "</style>",
                unsafe_allow_html=True,
            )
            cA, cB = st.columns(2)
            # ---- Apagar SELECIONADAS (com confirmação) ----
            with cA:
                if st.session_state.get("_conf_ap_sel"):
                    st.warning(f"Apagar {len(_sel)} selecionada(s)?")
                    if st.button("✅ Sim, apagar", key="ap_sel_sim", type="primary",
                                 width="stretch"):
                        for _d in _sel:
                            db.grade_apagar_lista(_map[_d], marca)
                        st.session_state.pop("_conf_ap_sel", None)
                        st.session_state.pop("_grade_sel", None)
                        st.session_state["_flash_grade"] = f"{len(_sel)} promoção(ões) apagada(s)!"
                        st.rerun()
                    if st.button("Cancelar", key="ap_sel_nao", width="stretch"):
                        st.session_state.pop("_conf_ap_sel", None)
                        st.rerun()
                elif st.button("🗑️ Apagar selecionadas", key="btn_ap_sel",
                               disabled=not _sel, width="stretch"):
                    st.session_state["_conf_ap_sel"] = True
                    st.rerun()
            # ---- Apagar TODAS (com confirmação) ----
            with cB:
                if st.session_state.get("_conf_ap_todas"):
                    st.warning(f"Apagar TODAS as {len(_listas)} promoções?")
                    if st.button("✅ Sim, apagar todas", key="ap_all_sim", type="primary",
                                 width="stretch"):
                        db.grade_apagar_todas(marca)
                        st.session_state.pop("_conf_ap_todas", None)
                        st.session_state.pop("_grade_sel", None)
                        st.session_state["_flash_grade"] = "Todas as promoções foram apagadas!"
                        st.rerun()
                    if st.button("Cancelar", key="ap_all_nao", width="stretch"):
                        st.session_state.pop("_conf_ap_todas", None)
                        st.rerun()
                elif st.button("🗑️ Apagar TODAS", key="btn_ap_all", width="stretch"):
                    st.session_state["_conf_ap_todas"] = True
                    st.rerun()


def pagina_promocoes(eh_admin: bool = False, marca: str = "natura") -> None:
    st.title("📑 Promoções da Grade")
    if st.session_state.get("_flash_grade"):     # aviso após apagar (sobrevive ao rerun)
        st.success(st.session_state.pop("_flash_grade"))
    if st.button("← Voltar ao calendário", key="voltar_calendario"):
        st.session_state.pop("_view", None)
        st.session_state.pop("_grade_sel", None)
        st.rerun()

    # --- Admin: subir uma grade nova ---------------------------------------
    if eh_admin:
        with st.expander("⬆️ Atualizar Grade (subir o Excel)", expanded=False):
            st.caption(
                "Suba o arquivo **.xlsm** da Grade de Ativação. O sistema lê as abas "
                "**LISTA_ visíveis** e substitui essas promoções (as que não vierem "
                "no arquivo ficam como histórico)."
            )
            arq = st.file_uploader("Arquivo da Grade", type=["xlsm", "xlsx"], key="up_grade")
            if arq is not None and st.button("Processar e salvar", type="primary", key="btn_proc"):
                try:
                    with st.spinner("Lendo a planilha e salvando no banco..."):
                        resumo = salvar_grade(arq.getvalue(), arq.name, marca)
                    if resumo:
                        st.success(f"✅ {len(resumo)} promoção(ões) atualizada(s)!")
                        for tipo, n in resumo:
                            st.caption(f"• {tipo}: {n} produtos")
                        st.session_state.pop("_grade_sel", None)
                    else:
                        st.warning("Nenhuma aba 'LISTA_' visível foi encontrada no arquivo.")
                except Exception as e:
                    _msg = str(e).lower()
                    if "zip" in _msg or "badzip" in _msg:
                        st.error(
                            "⚠️ Este arquivo parece **corrompido ou incompleto** — "
                            "geralmente é um **download que não terminou**, ou um arquivo "
                            "que está **só na nuvem** (OneDrive) e não baixou inteiro.\n\n"
                            "**O que fazer:** abra a pasta, confirme que o arquivo baixou "
                            "100% (ícone sem a nuvenzinha) ou **baixe de novo**, e tente "
                            "subir mais uma vez. 🙂"
                        )
                    else:
                        st.error(f"Não consegui ler o arquivo. Detalhe: {e}")

        # Painel de APAGAR promoções (seleção ou todas) — só admin.
        _painel_apagar_promocoes(marca)

    st.divider()

    sel = st.session_state.get("_grade_sel")
    if sel:
        _mostra_produtos(sel, marca)
    else:
        _mostra_lista_promocoes(marca)


def _mostra_lista_promocoes(marca: str = "natura") -> None:
    listas = db.grade_listar_listas(marca)
    if not listas:
        st.info("Nenhuma promoção cadastrada ainda. "
                + ("Suba uma Grade acima. ☝️" if True else ""))
        return

    # Cabeçalho destacado (Ciclo/período da grade) no espaço acima da busca. Como
    # as promoções do mesmo upload dividem o ciclo, mostra ele UMA vez, em destaque.
    _ciclos = sorted({(L.get("ciclo") or "").strip() for L in listas if L.get("ciclo")})
    _pers = sorted({(L.get("periodo") or "").strip() for L in listas if L.get("periodo")})
    if len(_ciclos) == 1:
        _hdr = f"🗂️ Ciclo {_ciclo_curto(_ciclos[0])}" + (
            f" · {_pers[0]}" if len(_pers) == 1 and _pers[0] else ""
        )
    else:
        _hdr = "🗂️ Grade de Promoções"
    # Cabeçalho: Ciclo à ESQUERDA + contagem de promoções à DIREITA, na MESMA faixa
    # azul (flex space-between). No celular, quebra em 2 linhas (flex-wrap).
    st.markdown(
        f"<div style='background:#EAF2FB;border:1px solid #1E88E5;border-radius:8px;"
        f"padding:0.55rem 0.95rem;margin:0.2rem 0 0.7rem;display:flex;"
        f"justify-content:space-between;align-items:center;flex-wrap:wrap;gap:0.4rem;'>"
        f"<span style='font-size:1.25rem;font-weight:800;color:#155FA0;'>{_hdr}</span>"
        f"<span style='font-size:1.05rem;font-weight:700;color:#155FA0;'>"
        f"📋 {len(listas)} ações disponíveis</span></div>",
        unsafe_allow_html=True,
    )

    busca = st.text_input("🔎 Buscar promoção", key="busca_lista").strip().lower()

    # Botões "ver ▸" (um por promoção, key="ver_<id>") com cara melhor: azul cheio +
    # negrito. Uso [class*='st-key-ver_'] p/ pegar todos de uma vez.
    st.markdown(
        "<style>"
        "[class*='st-key-ver_'] button{background:#EAF2FB !important;"
        "border:1px solid #1E88E5 !important;color:#155FA0 !important;"
        "font-weight:700 !important;border-radius:8px !important;}"
        "[class*='st-key-ver_'] button:hover{background:#D6E8FA !important;}"
        "</style>",
        unsafe_allow_html=True,
    )
    for L in listas:
        alvo = (str(L.get("lista_nome", "")) + " "
                + _tipo_amigavel(L.get("lista_nome", ""))).lower()
        if busca and busca not in alvo:
            continue
        c1, c2, c3 = st.columns([3, 2, 1.4])   # nome mais estreito: menos espaço vazio
        with c1:
            # Nome + período/LP num ÚNICO bloco, com line-height apertado, p/ a linha
            # ficar mais BAIXA (compacta). Nome completo calculado do lista_nome (já
            # vale p/ as listas atuais, sem re-subir a grade). Período = B1/C1; LP = B2.
            _sub = f"🗓️ {L.get('periodo_acao') or 'período —'}" + (
                f"  ·  🔗 LP: {L.get('link_lp')}" if L.get("link_lp") else ""
            )
            st.markdown(
                f"<div style='line-height:1.3;'>"
                f"<span style='font-weight:700;'>{_tipo_amigavel(L['lista_nome'])}</span><br>"
                f"<span style='color:#8a8a8a;font-size:0.85rem;'>{_sub}</span></div>",
                unsafe_allow_html=True,
            )
        c2.markdown(f"**{L.get('total_skus', 0)}** SKUs")
        if c3.button("ver ▸", key=f"ver_{L['id']}", width="stretch"):
            st.session_state._grade_sel = L["lista_nome"]
            st.rerun()
        # Divisória FINA (margem pequena) p/ os itens ficarem mais juntos/compactos.
        st.markdown(
            "<hr style='margin:0.35rem 0;border:none;border-top:1px solid #ECECEC;'>",
            unsafe_allow_html=True,
        )


def _mostra_produtos(lista_nome: str, marca: str = "natura") -> None:
    if st.button("← Voltar às promoções", key="voltar_prom"):
        st.session_state.pop("_grade_sel", None)
        st.rerun()

    L = next((x for x in db.grade_listar_listas(marca) if x["lista_nome"] == lista_nome), None)
    if L:
        st.subheader(_tipo_amigavel(lista_nome))
        # Descrição em faixa destacada (fonte maior + texto escuro + barra azul),
        # em vez do caption cinza pequeno de antes.
        _info = (
            f"<b>{L.get('total_skus', 0)} produtos</b> · 🗓️ "
            f"{L.get('periodo_acao') or 'período —'}"
            + (f" · 🔗 LP: {L.get('link_lp')}" if L.get("link_lp") else "")
        )
        st.markdown(
            f"<div style='background:#F0F2F6;border-left:4px solid #1E88E5;"
            f"border-radius:6px;padding:0.5rem 0.9rem;margin:0.1rem 0 0.7rem;"
            f"font-size:1.02rem;color:#2A2A3C;'>{_info}</div>",
            unsafe_allow_html=True,
        )

    produtos = db.grade_listar_produtos(lista_nome, marca)
    if not produtos:
        st.info("Esta promoção está sem produtos.")
        return

    df = pd.DataFrame(produtos)
    # Só as colunas escolhidas, na ordem certa, com rótulos amigáveis.
    df = df[[c for c in _COLS if c in df.columns]].rename(columns=_ROTULOS)

    busca = st.text_input("🔎 Buscar produto (SKU ou descrição)", key="busca_prod").strip().lower()
    if busca:
        m = df["SKU"].astype(str).str.lower().str.contains(busca, na=False) | \
            df["Descrição"].astype(str).str.lower().str.contains(busca, na=False)
        df = df[m]

    # Exibição: DE/POR com 2 casas (164,90); Desconto em % sem casas (17%). O desconto
    # vem como FRAÇÃO (0.1672 = 16,72%), então ×100 p/ o formato "%.0f%%" mostrar "17%".
    _show = df.copy()
    _cfg = {}
    for _c in ("DE", "POR"):
        if _c in _show.columns:
            _show[_c] = pd.to_numeric(_show[_c], errors="coerce")
            _cfg[_c] = st.column_config.NumberColumn(_c, format="%.2f")
    if "Desconto" in _show.columns:
        _show["Desconto"] = pd.to_numeric(_show["Desconto"], errors="coerce") * 100
        _cfg["Desconto"] = st.column_config.NumberColumn("Desconto", format="%.0f%%")
    st.dataframe(_show, width="stretch", hide_index=True, column_config=_cfg)
    _nome_arq = _tipo_amigavel(lista_nome) or "promocao"
    _c_csv, _c_xls, _ = st.columns([1, 1, 3])
    with _c_csv:
        st.download_button(
            "⬇️ Baixar CSV",
            data=df.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"{_nome_arq}.csv",
            mime="text/csv",
            key="dl_csv_grade",
            help="Abre no Google Sheets (Arquivo → Importar) e em qualquer planilha.",
        )
    with _c_xls:
        st.download_button(
            "⬇️ Baixar Excel",
            data=_excel_bytes(df),
            file_name=f"{_nome_arq}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_grade",
        )
