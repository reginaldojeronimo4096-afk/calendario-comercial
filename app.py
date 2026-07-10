# -*- coding: utf-8 -*-
"""
Calendário de Ações e Promoções — formato linha do tempo (timeline / Gantt).
Preencha os campos, e a ação é desenhada como um bloco colorido no período certo.
Dados guardados no Supabase (banco online) — ver db.py. Login e papéis em auth.py.
"""

import io
import os
import textwrap
from datetime import date, datetime, timedelta

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import auth
import db
import grade

# ----------------------------------------------------------------------------
# Configuração e constantes
# ----------------------------------------------------------------------------
ARQUIVO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "acoes.csv")

COLUNAS = ["Ação", "Categoria", "Início", "Fim", "Cor", "Detalhes"]

# Portfólio do Ciclo: faixa amarela no topo (acima das semanas) com o texto
# digitado (ex.: "C09 e C10"), cobrindo o período escolhido. Salvo à parte.
ARQUIVO_CICLOS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ciclos.csv")
COLUNAS_CICLOS = ["Ciclo", "Início", "Fim"]

# Faixas (lanes) FIXAS — aparecem SEMPRE no calendário, nesta ordem exata,
# mesmo que o mês não tenha ação nelas (a linha fica vazia, mas o título no lugar).
# Espelham as seções da coluna A/B da planilha "Grade Comercial".
CATEGORIAS_PADRAO = [
    "DESTAQUE DA COMUNICAÇÃO",
    "PRESENTES",
    "PROGRESSIVO",
    "MONTE SEU RITUAL",
    "KITS",
    "PRIMEIRA COMPRA",
    "FRETE",
    "DE/POR CICLO",
    "FLASHSALE / EFEMÉRIDES / DE-POR EXCLUSIVO",
    "PROMO CICLO ABERTO",
    "OFERTA RELÂMPAGO",
    "EXTRAS",
    "SOBRAS",
    "CUPOM SITE",
    "BRINDE",
    "AMOSTRAS",
    "CALENDÁRIO CRM",
]

# A ordem fixa das faixas no calendário é exatamente a lista acima.
FAIXAS_FIXAS = list(CATEGORIAS_PADRAO)

# A 1ª faixa ("DESTAQUE DA COMUNICAÇÃO") fica GRUDADA no cabeçalho (não rola junto
# com as demais) — desenhada num gráfico próprio (fig_destaque) na zona fixa do
# topo. FAIXAS_CORPO são as faixas que rolam no corpo do calendário.
FAIXA_FIXA_TOPO = FAIXAS_FIXAS[0]
FAIXAS_CORPO = [c for c in FAIXAS_FIXAS if c != FAIXA_FIXA_TOPO]

# Quebras de linha MANUAIS para rótulos específicos (mantêm termos juntos em vez
# de deixar a quebra automática separar palavras que devem ficar na mesma linha).
QUEBRA_MANUAL = {
    "FLASHSALE / EFEMÉRIDES / DE-POR EXCLUSIVO":
        "FLASHSALE /<br>EFEMÉRIDES /<br>DE-POR EXCLUSIVO",
}

# Equivalências: nomes de faixa antigos (já salvos no acoes.csv) -> faixa nova.
# Assim as ações antigas caem na faixa certa sem precisar reeditar o CSV.
ALIAS_FAIXAS = {
    # Nomes antigos / variações já salvos no acoes.csv -> faixa nova (canônica).
    "Destaque Comunicação": "DESTAQUE DA COMUNICAÇÃO",
    "Kits": "KITS",
    "De/Por Ciclo": "DE/POR CICLO",
    "Oferta Relâmpago": "OFERTA RELÂMPAGO",
    "Brinde": "BRINDE",
    "Extra": "EXTRAS",
    "Extras": "EXTRAS",
    "Cupom Site": "CUPOM SITE",
    "Cupom SITE": "CUPOM SITE",
    "Flashsale / Efemérides / De/Por": "FLASHSALE / EFEMÉRIDES / DE-POR EXCLUSIVO",
    "Flashsale / Efemérides": "FLASHSALE / EFEMÉRIDES / DE-POR EXCLUSIVO",
}


def normaliza_categoria(cat) -> str:
    """Tira espaços/quebras de linha e converte nomes antigos p/ a faixa nova."""
    c = str(cat).strip()
    return ALIAS_FAIXAS.get(c, c)

# Cores das faixas do Portfólio do Ciclo — alternadas por ordem de início,
# para distinguir ciclos diferentes no mesmo mês (amarelo, azul claro, ...).
CORES_CICLO = [
    {"fill": "#F8E71C", "borda": "#E1C400", "texto": "#5A4B00"},  # amarelo
    {"fill": "#85C1E9", "borda": "#2E86C1", "texto": "#0E3A5A"},  # azul claro
]

CORES_SUGERIDAS = {
    "Vermelho": "#E8463A",
    "Laranja": "#F5A623",
    "Amarelo": "#F8E71C",
    "Verde": "#3FB950",
    "Azul": "#2E86DE",
    "Roxo": "#9B59B6",
    "Rosa": "#E84393",
    "Cinza": "#95A5A6",
}

# Paleta de cores da ação — mostrada como quadradinhos clicáveis.
PALETA_CORES = [
    "#FF0000", "#990000", "#F4CCCC", "#FFC000", "#0000FF", "#000000",
    "#7F7F7F", "#BFBFBF", "#CC33FF", "#38761D", "#6AA84F", "#E46C0A",
    "#FFFFFF", "#93D050", "#A9D18E", "#E06666", "#FF4343", "#61CBF3",
    "#83CCEB", "#741B47", "#00FFFF", "#F9CB9C",
]


def _contraste_texto(hex_cor: str) -> str:
    """Preto ou branco conforme a luminância da cor — mantém o texto legível
    sobre qualquer cor de fundo."""
    h = str(hex_cor).lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    try:
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except (ValueError, IndexError):
        return "#000000"
    return "#000000" if (0.299 * r + 0.587 * g + 0.114 * b) > 150 else "#FFFFFF"

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

st.set_page_config(page_title="Calendário de Ações", page_icon="📅", layout="wide")

# Esconde só o botão "Implantar" (mantém o menu ⋮ com tema claro/escuro) e
# reduz o espaço em branco acima do cabeçalho, deixando o título mais no topo.
st.markdown(
    """
    <style>
      [data-testid="stDeployButton"],
      [data-testid="stAppDeployButton"] { display: none !important; }
      [data-testid="stHeader"] { background: transparent; }
      /* padding-top: título mais no topo. padding-bottom: encolhe o grande vazio
         que o Streamlit reserva no rodapé (~10rem). Esse vazio deixava rolar a
         página "além do fim", empurrando a linha SEMANA do cabeçalho para trás
         da barra do navegador (sobrava só as datas). Mantido pequeno (2.5rem) para
         dar um respiro no rodapé SEM voltar aquele problema de rolar demais. */
      .block-container { padding-top: 0.5rem; padding-bottom: 2.5rem; }
      /* Título do calendário numa linha só, mesmo com "Dezembro" */
      h1 { font-size: 2.1rem !important; white-space: nowrap; }
    </style>
    """,
    unsafe_allow_html=True,
)


# ----------------------------------------------------------------------------
# Login / controle de acesso (ANTES de tudo do calendário)
# ----------------------------------------------------------------------------
# Cria os admins iniciais (se não existirem) e exige login. Enquanto ninguém
# estiver logado, mostra só a tela de login e para aqui (st.stop). Depois de
# logado, PAPEL define o que a pessoa pode fazer.
auth.inicializar()
if not auth.esta_logado():
    auth.tela_login()
    st.stop()

_U = auth.usuario_atual()
PAPEL = _U["papel"]
PODE_EDITAR = PAPEL in ("admin", "editor")   # admin/editor mexem no calendário
EH_ADMIN = PAPEL == "admin"                  # só admin gerencia usuários

# Barra lateral: quem está logado + Promoções + botão de sair.
with st.sidebar:
    st.markdown(f"### 👤 {_U.get('nome') or _U['usuario']}")
    st.caption(f"Papel: **{PAPEL}**")
    if st.button("🚪 Sair", use_container_width=True):
        auth.sair()
    st.divider()
    # Abre a tela de consulta das promoções (Grade de Ativação) — para todos.
    if st.button("📑 Promoções", use_container_width=True):
        st.session_state["_view"] = "promocoes"
        st.session_state.pop("_grade_sel", None)
        st.rerun()

# Se estiver no modo "Promoções", mostra essa página e PARA (não desenha o
# calendário). O botão "Voltar ao calendário" limpa o _view e volta ao normal.
if st.session_state.get("_view") == "promocoes":
    grade.pagina_promocoes(EH_ADMIN)
    st.stop()


# ----------------------------------------------------------------------------
# Traduz o seletor de data para português (dias da semana e meses).
# O tradutor automático do navegador traduz "Tu"/"We" (Tue/Wed) como os
# pronomes "Você"/"Nós". Este script desliga a tradução da página e força
# dias (por POSIÇÃO) e meses (por nome) em português — resiste à tradução.
# ----------------------------------------------------------------------------
def _traduz_calendario() -> None:
    st.components.v1.html(
        """
        <script>
        const PT = ['Dom','Seg','Ter','Qua','Qui','Sex','Sáb'];
        const RE = /^(Su|Mo|Tu|We|Th|Fr|Sa|Você|Nós|Dom|Seg|Ter|Qua|Qui|Sex|S[áa]b)$/;
        const MES = {January:'Janeiro', February:'Fevereiro', March:'Março',
          April:'Abril', May:'Maio', June:'Junho', July:'Julho', August:'Agosto',
          September:'Setembro', October:'Outubro', November:'Novembro',
          December:'Dezembro'};
        function fixDias(doc) {
          const cal = doc.querySelector('[data-baseweb="calendar"]');
          if (!cal) return;
          const cells = Array.from(cal.querySelectorAll('*')).filter(
            e => e.children.length === 0 && RE.test((e.textContent || '').trim())
          );
          for (let i = 0; i < Math.min(7, cells.length); i++) {
            if (cells[i].textContent.trim() !== PT[i]) cells[i].textContent = PT[i];
          }
        }
        function fixMeses(doc) {
          // Cabeçalho do calendário + menu suspenso de meses (popover/listbox).
          // Varre os NÓS DE TEXTO p/ pegar também o mês no cabeçalho, que fica
          // num botão junto com a setinha (não é um elemento "folha").
          const escopos = doc.querySelectorAll(
            '[data-baseweb="calendar"], [data-baseweb="popover"], [role="listbox"]');
          escopos.forEach(escopo => {
            const walker = doc.createTreeWalker(escopo, NodeFilter.SHOW_TEXT);
            let no;
            while ((no = walker.nextNode())) {
              const t = (no.nodeValue || '').trim();
              if (MES[t]) no.nodeValue = MES[t];
            }
          });
        }
        function fixPlaceholders(doc) {
          // Texto interno "Insira a Data" nos campos de data (que começam vazios).
          doc.querySelectorAll('[data-testid="stDateInput"] input').forEach(
            function (inp) {
              if (inp.placeholder !== 'Insira a Data') inp.placeholder = 'Insira a Data';
            }
          );
        }
        function fix() {
          try {
            const doc = window.parent.document;
            doc.documentElement.setAttribute('translate', 'no');
            doc.documentElement.classList.add('notranslate');
            fixDias(doc);
            fixMeses(doc);
            fixPlaceholders(doc);
          } catch (e) {}
        }
        try { new MutationObserver(fix).observe(window.parent.document.body,
              {childList: true, subtree: true}); } catch (e) {}
        setInterval(fix, 300);
        fix();
        </script>
        """,
        height=0,
    )


_traduz_calendario()


# ----------------------------------------------------------------------------
# Persistência (Supabase — ver db.py)
# ----------------------------------------------------------------------------
# As colunas do app têm acento ("Ação", "Início"); no banco são sem acento
# ("acao", "inicio"). Estes mapas traduzem de um lado para o outro.
_MAPA_ACOES = {
    "acao": "Ação", "categoria": "Categoria", "inicio": "Início",
    "fim": "Fim", "cor": "Cor", "detalhes": "Detalhes",
}
_MAPA_CICLOS = {"ciclo": "Ciclo", "inicio": "Início", "fim": "Fim"}


def _data_iso(v):
    """Converte uma data (ou texto) para 'AAAA-MM-DD' — ou None se vazia."""
    try:
        if v is None or pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, str):
        return v[:10] if v.strip() else None
    try:
        return v.isoformat()[:10]  # date/Timestamp -> 'AAAA-MM-DD' (sem a hora)
    except (AttributeError, TypeError, ValueError):
        return None


def _para_texto(v):
    """Texto seguro (nunca None/NaN) para gravar no banco.
    NOME ÚNICO de propósito: NÃO usar '_texto' — esse nome é reaproveitado como
    variável no desenho dos ciclos (lá embaixo), o que atropelaria esta função."""
    try:
        if v is None or pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return str(v)


def carregar() -> pd.DataFrame:
    linhas = db.listar_acoes()
    if not linhas:
        return pd.DataFrame(columns=COLUNAS)
    df = pd.DataFrame(linhas).rename(columns=_MAPA_ACOES)
    for c in COLUNAS:
        if c not in df.columns:
            df[c] = ""
    df = df[COLUNAS]
    df["Início"] = pd.to_datetime(df["Início"], errors="coerce").dt.date
    df["Fim"] = pd.to_datetime(df["Fim"], errors="coerce").dt.date
    return df


def salvar(df: pd.DataFrame) -> None:
    # to_dict("records") converte para dicts Python simples (valores nativos),
    # evitando peculiaridades do acesso linha-a-linha do pandas (que davam
    # TypeError ao juntar a ação nova, no Python novo da hospedagem).
    registros = [
        {
            "acao": _para_texto(rec.get("Ação")),
            "categoria": _para_texto(rec.get("Categoria")),
            "inicio": _data_iso(rec.get("Início")),
            "fim": _data_iso(rec.get("Fim")),
            "cor": _para_texto(rec.get("Cor")),
            "detalhes": _para_texto(rec.get("Detalhes")),
        }
        for rec in df.to_dict(orient="records")
    ]
    db.substituir_acoes(registros)


def carregar_ciclos() -> pd.DataFrame:
    linhas = db.listar_ciclos()
    if not linhas:
        return pd.DataFrame(columns=COLUNAS_CICLOS)
    df = pd.DataFrame(linhas).rename(columns=_MAPA_CICLOS)
    for c in COLUNAS_CICLOS:
        if c not in df.columns:
            df[c] = ""
    df = df[COLUNAS_CICLOS]
    df["Início"] = pd.to_datetime(df["Início"], errors="coerce").dt.date
    df["Fim"] = pd.to_datetime(df["Fim"], errors="coerce").dt.date
    return df


def salvar_ciclos(df: pd.DataFrame) -> None:
    registros = [
        {
            "ciclo": _para_texto(rec.get("Ciclo")),
            "inicio": _data_iso(rec.get("Início")),
            "fim": _data_iso(rec.get("Fim")),
        }
        for rec in df.to_dict(orient="records")
    ]
    db.substituir_ciclos(registros)


# ----------------------------------------------------------------------------
# Estado da sessão
# ----------------------------------------------------------------------------
if "df" not in st.session_state:
    st.session_state.df = carregar()

if "ciclos_df" not in st.session_state:
    st.session_state.ciclos_df = carregar_ciclos()


# ----------------------------------------------------------------------------
# Cabeçalho
# ----------------------------------------------------------------------------
hoje = date.today()
col_titulo, col_mes, col_ano = st.columns([7, 1, 1])
with col_mes:
    mes_num = st.selectbox(
        "Mês",
        options=list(range(1, 13)),
        format_func=lambda m: MESES_PT[m - 1],
        index=hoje.month - 1,
    )
with col_ano:
    ano = int(
        st.number_input("Ano", min_value=2024, max_value=2032, value=hoje.year, step=1)
    )
with col_titulo:
    st.title(f"📅 Calendário da Grade Comercial de {MESES_PT[mes_num - 1]}")

# Mensagem de feedback que sobrevive ao recarregamento (ex.: após limpar tudo)
if st.session_state.get("flash"):
    st.success(st.session_state.pop("flash"))


# ----------------------------------------------------------------------------
# Formulário: adicionar nova ação  (agora em POP-UP / modal)
# ----------------------------------------------------------------------------
# Antes o formulário ficava num expander fixo ACIMA do calendário, empurrando o
# calendário para baixo. Agora ele abre num pop-up (st.dialog) acionado por um
# botão — assim o CALENDÁRIO ocupa o topo da página. O corpo do formulário é o
# MESMO de antes; só mudou a "moldura" (de expander para função de diálogo).
# Recebe ano/mes_num por parâmetro: como o diálogo é um "fragment", ao clicar
# nos botões da paleta só ele re-executa, então precisa desses valores fixados.
@st.dialog("➕ Adicionar nova ação", width="large")
def dialog_adicionar(ano, mes_num):
    # Cor da ação: os defaults ficam aqui para o salvamento ter sempre um valor.
    if "cor_escolhida" not in st.session_state:
        st.session_state.cor_escolhida = PALETA_CORES[0]
    if "cor_tom" not in st.session_state:
        st.session_state.cor_tom = PALETA_CORES[0]

    # Limpa os campos após um cadastro bem-sucedido: cada widget usa uma chave
    # com sufixo "_{_n}"; ao salvar, incrementamos 'form_nonce', então na próxima
    # execução os campos são widgets NOVOS (vazios). Método à prova de falhas,
    # sem depender de apagar a chave (que o Streamlit nem sempre reseta).
    st.session_state.setdefault("form_nonce", 0)
    _n = st.session_state.form_nonce

    # Frase de instrução logo abaixo do título do pop-up (antes ficava embaixo do
    # título da página; movida para cá para orientar quem abre o formulário).
    st.caption("Preencha os campos abaixo para adicionar ao Calendário")

    with st.container():
        # Limites do seletor de data = mês selecionado no topo. Assim o popup do
        # calendário já abre NO MÊS que está sendo montado (não no mês atual) e só
        # permite datas daquele mês. '_mesnonce' entra na chave das datas para o
        # campo resetar ao trocar de mês (senão a data ficaria fora do limite).
        _mes_ini = date(ano, mes_num, 1)
        _prox_mes = date(ano + 1, 1, 1) if mes_num == 12 else date(ano, mes_num + 1, 1)
        _mes_fim = _prox_mes - timedelta(days=1)
        _mesnonce = f"{ano}{mes_num:02d}"

        # Linha 0: Portfólio do Ciclo + período do portfólio (lado a lado)
        col_portfolio, col_portf_ini, col_portf_fim = st.columns([2, 1, 1])
        with col_portfolio:
            portfolio_ciclo = st.text_input(
                "Portfólio do Ciclo",
                placeholder="Digite os Ciclos do período",
                key=f"portfolio_ciclo_{_n}",
            )
        with col_portf_ini:
            # Datas ÚNICAS: valem para o Ciclo (se preenchido) e/ou para a Ação.
            data_inicio = st.date_input(
                "Data de início *",
                value=None,
                format="DD/MM/YYYY",
                min_value=_mes_ini,
                max_value=_mes_fim,
                key=f"data_inicio_{_mesnonce}_{_n}",
            )
        with col_portf_fim:
            data_fim = st.date_input(
                "Data final *",
                value=None,
                format="DD/MM/YYYY",
                min_value=_mes_ini,
                max_value=_mes_fim,
                key=f"data_fim_{_mesnonce}_{_n}",
            )

        # Linha 1: nome da ação + faixa + faixa nova (lado a lado)
        col_nome, col_faixa, col_faixa_nova = st.columns(3)
        with col_nome:
            nome = st.text_input(
                "Nome da ação *", placeholder="Ex: Dia do Meio Ambiente",
                key=f"nome_{_n}",
            )
        with col_faixa:
            categoria = st.selectbox(
                "Faixa (lado esquerdo do calendário)",
                options=CATEGORIAS_PADRAO,
                index=0,
                key=f"categoria_{_n}",
            )
        with col_faixa_nova:
            categoria_nova = st.text_input(
                "...ou digite uma faixa nova",
                placeholder="Ex: PRESENTES",
                key=f"categoria_nova_{_n}",
            )

        # Linha 2: descrição da ação em largura total (as datas agora ficam na
        # linha de cima e valem tanto para o Ciclo quanto para a Ação).
        detalhes = st.text_area(
            "Descrição da ação",
            placeholder="Ex: 5% OFF em todo o site + cupom MEIOAMBIENTE",
            height=120,
            key=f"detalhes_{_n}",
        )

    # --- Paleta de cores (discreta), logo acima do botão de salvar ----------
    # Fica fora de qualquer form (botões não disparam dentro de um formulário);
    # a cor escolhida vai para st.session_state e é lida ao salvar.
    st.markdown(
        "<div style='color:#4F4F4F;font-size:0.875rem;margin:0 0 0.35rem;'>"
        "🎨 Cor da ação — Selecione a cor da linha</div>",
        unsafe_allow_html=True,
    )
    # "Ajustar tom" à ESQUERDA; a paleta ocupa a direita (retângulos puxados p/ lá).
    # IMPORTANTE: a paleta é preenchida ANTES do color_picker no código porque os
    # botões definem st.session_state.cor_tom — e isso precisa ocorrer antes do
    # widget de key "cor_tom" existir (senão o Streamlit lança erro). A posição
    # visual (tom à esquerda) vem da ORDEM DAS COLUNAS, não da ordem do código.
    col_tom, col_paleta = st.columns([1, 9], vertical_alignment="center")
    with col_paleta:
        POR_LINHA = 11
        for ini in range(0, len(PALETA_CORES), POR_LINHA):
            cols = st.columns(POR_LINHA)
            for j, hexcor in enumerate(PALETA_CORES[ini:ini + POR_LINHA]):
                i = ini + j
                if cols[j].button(" ", key=f"cor_btn_{i}", use_container_width=True,
                                  help=hexcor):
                    st.session_state.cor_escolhida = hexcor
                    st.session_state.cor_tom = hexcor
    with col_tom:
        st.color_picker("Ajustar tom", key="cor_tom")

    _css = ["<style>"]
    for i, hexcor in enumerate(PALETA_CORES):
        _css.append(
            f".st-key-cor_btn_{i} button {{"
            f"background-color:{hexcor} !important;"
            f"border:1px solid #C4C4C4 !important;border-radius:4px !important;"
            f"height:18px;min-height:18px;padding:0 !important;}}"
            f".st-key-cor_btn_{i} button:hover{{"
            f"background-color:{hexcor} !important;border-color:#555 !important;}}"
        )
    if st.session_state.cor_escolhida in PALETA_CORES:
        _sel = PALETA_CORES.index(st.session_state.cor_escolhida)
        _css.append(
            f".st-key-cor_btn_{_sel} button {{"
            f"border:2px solid #111 !important;"
            f"box-shadow:0 0 0 1px #fff inset !important;}}"
        )
    _css.append("</style>")
    st.markdown("\n".join(_css), unsafe_allow_html=True)

    # Botão de salvar — por último, abaixo de tudo.
    enviar = st.button("Adicionar ao calendário", type="primary")

    if enviar:
        cat_final = categoria_nova.strip() if categoria_nova.strip() else categoria
        tem_ciclo = bool(portfolio_ciclo.strip())
        tem_acao = bool(nome.strip())

        if not tem_ciclo and not tem_acao:
            st.error("Preencha o nome da ação ou um Portfólio do Ciclo. 🙂")
        elif data_inicio is None or data_fim is None:
            st.error("Informe as datas (Data de início e Data final).")
        elif data_fim < data_inicio:
            st.error("A Data final não pode ser antes da Data de início.")
        else:
            mensagens = []
            # Portfólio do Ciclo -> faixa amarela no topo do calendário.
            # Se o novo período se sobrepõe a um ciclo já existente, ele
            # SUBSTITUI o anterior (evita faixas duplicadas no mesmo trecho).
            if tem_ciclo:
                cdf = st.session_state.ciclos_df
                if not cdf.empty:
                    sobrepoe = (cdf["Início"] <= data_fim) & (
                        cdf["Fim"] >= data_inicio
                    )
                    cdf = cdf[~sobrepoe]
                novo_ciclo = {
                    "Ciclo": portfolio_ciclo.strip(),
                    "Início": data_inicio,
                    "Fim": data_fim,
                }
                st.session_state.ciclos_df = pd.concat(
                    [cdf, pd.DataFrame([novo_ciclo])], ignore_index=True
                )
                salvar_ciclos(st.session_state.ciclos_df)
                mensagens.append(f"Ciclo “{portfolio_ciclo.strip()}” adicionado!")
            # Ação -> bloco colorido na linha do tempo
            if tem_acao:
                nova = {
                    "Ação": nome.strip(),
                    "Categoria": cat_final,
                    "Início": data_inicio,
                    "Fim": data_fim,
                    "Cor": st.session_state.get("cor_tom", PALETA_CORES[0]),
                    "Detalhes": detalhes.strip(),
                }
                st.session_state.df = pd.concat(
                    [st.session_state.df, pd.DataFrame([nova])], ignore_index=True
                )
                salvar(st.session_state.df)
                mensagens.append(f"Ação “{nome.strip()}” adicionada!")
            # Incrementar o nonce recria os campos vazios na próxima abertura do
            # pop-up (limpa o formulário). O st.rerun() fecha o diálogo e redesenha
            # o calendário já com a nova ação/ciclo cadastrados.
            st.session_state.form_nonce += 1
            st.rerun()


# ----------------------------------------------------------------------------
# Botão que abre o pop-up "Adicionar nova ação"
# ----------------------------------------------------------------------------
# O formulário agora vive na função de diálogo definida acima; aqui fica só o
# botão que o abre. Fica logo abaixo do título e ACIMA do calendário, para o
# calendário ocupar o topo da página (menos rolagem até ele).
# Botões conforme o papel: admin/editor veem "Adicionar"; só admin vê "Gerenciar
# Usuários" (botão verde). Leitor não vê nenhum dos dois (só olha o calendário).
if PODE_EDITAR or EH_ADMIN:
    _bt1, _bt2, _bt3 = st.columns([2, 2, 6])
    if PODE_EDITAR:
        with _bt1:
            if st.button("➕ Adicionar nova ação", type="primary",
                         use_container_width=True):
                dialog_adicionar(ano, mes_num)
    if EH_ADMIN:
        with _bt2:
            if st.button("👥 Gerenciar Usuários", key="btn_gerenciar",
                         use_container_width=True):
                auth.dialog_gerenciar_usuarios()
    # Deixa o botão "Gerenciar Usuários" verde (o de adicionar continua vermelho).
    st.markdown(
        "<style>.st-key-btn_gerenciar button{background:#21A038!important;"
        "border:0!important;color:#fff!important;font-weight:700!important;}"
        ".st-key-btn_gerenciar button:hover{background:#1B8730!important;}</style>",
        unsafe_allow_html=True,
    )


# ----------------------------------------------------------------------------
# Filtro por mês (opcional)
# ----------------------------------------------------------------------------
df = st.session_state.df.copy()
df = df.dropna(subset=["Início", "Fim"])

# Recorte do mês escolhido no seletor do topo
sel_inicio = date(ano, mes_num, 1)
sel_fim_excl = date(ano + 1, 1, 1) if mes_num == 12 else date(ano, mes_num + 1, 1)
sel_fim = sel_fim_excl - timedelta(days=1)

# Mostra as ações que se sobrepõem ao mês (mesmo que comecem/terminem fora dele)
if df.empty:
    df_view = df.copy()
else:
    df_view = df[(df["Início"] <= sel_fim) & (df["Fim"] >= sel_inicio)].copy()


# ----------------------------------------------------------------------------
# Gráfico: linha do tempo (timeline / Gantt)
# ----------------------------------------------------------------------------
st.subheader("🗓️ Calendário")

# Mês sem ações: mostra a grade vazia do mês (bloco transparente "fantasma"),
# em vez de erro, e avisa que ainda não há ações.
if df_view.empty:
    st.info(
        f"Nenhuma ação em {MESES_PT[mes_num - 1]} de {ano} ainda. "
        "Adicione no formulário acima. ☝️"
    )
    df_view = pd.DataFrame([{
        "Ação": "", "Categoria": FAIXAS_FIXAS[0], "Início": sel_inicio, "Fim": sel_inicio,
        "Cor": "rgba(0,0,0,0)", "Detalhes": "",
    }])

plot_df = df_view.copy()
# Normaliza o nome da faixa (tira espaços/quebras e aplica as equivalências
# de nomes antigos), p/ casar com a lista fixa FAIXAS_FIXAS.
plot_df["Categoria"] = plot_df["Categoria"].map(normaliza_categoria)
# Borda esquerda no INÍCIO do dia-rótulo e direita no FIM do dia-rótulo (cada
# rótulo ocupa meio-dia para cada lado do seu marco) — alinhamento idêntico ao
# da faixa do Portfólio do Ciclo. Vale para todas as barras.
meia = timedelta(hours=12)
plot_df["Início_plot"] = pd.to_datetime(plot_df["Início"]) - meia
plot_df["Fim_plot"] = pd.to_datetime(plot_df["Fim"]) + meia
def _rotulo_barra(texto: str) -> str:
    """Quebra o nome da ação em 2 linhas equilibradas (corte no espaço mais
    próximo do meio). Sem espaço, retorna como está. A DECISÃO de quebrar fica
    na lógica de largura abaixo (só quebra quando o nome não cabe em 1 linha)."""
    t = str(texto).strip()
    if " " not in t:
        return t
    meio = len(t) // 2
    esq = t.rfind(" ", 0, meio)   # espaço antes do meio
    dir_ = t.find(" ", meio)      # espaço depois do meio
    cand = [p for p in (esq, dir_) if p != -1]
    if not cand:
        return t
    corte = min(cand, key=lambda p: abs(p - meio))  # o mais próximo do meio
    return t[:corte] + "<br>" + t[corte + 1:]


# Largura visível de cada ação (fração do mês) -> estimativa de quantos
# caracteres cabem em 1 linha DENTRO da barra. Só quebra em 2 linhas quando o
# nome não cabe; barras largas (Chronos/Ekos/Lumina, full-width, etc.) ficam em
# 1 linha. CHARS_LARGURA_CHEIA = capacidade aprox. na largura TOTAL do calendário
# — se estiver quebrando cedo/tarde demais, é só ajustar este número.
CHARS_LARGURA_CHEIA = 75
_ini_vis = pd.to_datetime(plot_df["Início"]).clip(lower=pd.Timestamp(sel_inicio))
_fim_vis = pd.to_datetime(plot_df["Fim"]).clip(upper=pd.Timestamp(sel_fim))
_dias_vis = (_fim_vis - _ini_vis).dt.days + 1
_total_dias = max((pd.Timestamp(sel_fim) - pd.Timestamp(sel_inicio)).days + 1, 1)
_chars_cabem = (_dias_vis / _total_dias) * CHARS_LARGURA_CHEIA

# Rótulo em negrito; quebra só quando o nome não cabe em 1 linha naquela largura.
plot_df["Rótulo"] = [
    "<b>" + (str(a) if len(str(a)) <= cabem else _rotulo_barra(str(a))) + "</b>"
    for a, cabem in zip(plot_df["Ação"], _chars_cabem)
]
# Datas formatadas em DD/MM/AAAA (padrão BR) para o tooltip.
plot_df["Início_fmt"] = pd.to_datetime(plot_df["Início"]).dt.strftime("%d/%m/%Y")
plot_df["Fim_fmt"] = pd.to_datetime(plot_df["Fim"]).dt.strftime("%d/%m/%Y")

# ---------------------------------------------------------------------------
# Faixas (lanes) à esquerda — como a coluna A/B do Excel.
# As faixas seguem a ordem das categorias padrão; faixas novas (digitadas
# por você) entram logo depois, na ordem em que aparecem.
# ---------------------------------------------------------------------------
plot_df = plot_df.sort_values("Início_plot").reset_index(drop=True)

# Separa a faixa fixa do topo ("DESTAQUE DA COMUNICAÇÃO") — desenhada à parte,
# grudada no cabeçalho — do restante das faixas (o corpo que rola).
def _linha_vazia(cat):
    """Linha transparente (1 dia) só p/ a faixa aparecer VAZIA quando o mês não
    tem ação nela — mantém a MESMA estrutura de colunas já derivadas no plot_df."""
    d0 = pd.Timestamp(sel_inicio)
    return {
        "Ação": "", "Categoria": cat, "Início": sel_inicio, "Fim": sel_inicio,
        "Cor": "rgba(0,0,0,0)", "Detalhes": "",
        "Início_plot": d0 - meia, "Fim_plot": d0 + meia,
        "Rótulo": "", "Início_fmt": d0.strftime("%d/%m/%Y"),
        "Fim_fmt": d0.strftime("%d/%m/%Y"),
    }

topo_df = plot_df[plot_df["Categoria"] == FAIXA_FIXA_TOPO].copy().reset_index(drop=True)
plot_df = plot_df[plot_df["Categoria"] != FAIXA_FIXA_TOPO].copy().reset_index(drop=True)
if topo_df.empty:                     # mês sem comunicação -> faixa fixa vazia
    topo_df = pd.DataFrame([_linha_vazia(FAIXA_FIXA_TOPO)])
if plot_df.empty:                     # corpo sem nada -> ao menos 1 faixa vazia
    plot_df = pd.DataFrame([_linha_vazia(FAIXAS_CORPO[0])])

cats_presentes = list(dict.fromkeys(plot_df["Categoria"].tolist()))
# Faixas fixas SEMPRE visíveis (na ordem definida) + faixas extras que
# porventura existam nos dados e não estejam na lista fixa (vão para o fim).
# FAIXAS_CORPO já EXCLUI a faixa fixa do topo (que sai no fig_destaque).
extras = [c for c in cats_presentes if c not in FAIXAS_FIXAS and str(c).strip()]
ordem_cats = FAIXAS_CORPO + extras

# Empilha ações que se SOBREPÕEM na mesma faixa em "sub-linhas", para nenhuma
# ficar escondida atrás da outra (resolve o caso de blocos no mesmo período).
nivel = {}
for cat in ordem_cats:
    fins_por_nivel = []  # fim da última ação alocada em cada sub-linha
    for i in plot_df.index[plot_df["Categoria"] == cat]:
        inicio_i, fim_i = plot_df.at[i, "Início_plot"], plot_df.at[i, "Fim_plot"]
        colocado = False
        for n, fim_n in enumerate(fins_por_nivel):
            if inicio_i >= fim_n:              # não sobrepõe -> reaproveita a sub-linha
                fins_por_nivel[n] = fim_i
                nivel[i] = n
                colocado = True
                break
        if not colocado:                       # sobrepõe -> abre nova sub-linha
            fins_por_nivel.append(fim_i)
            nivel[i] = len(fins_por_nivel) - 1

n_niveis = {}
for cat in ordem_cats:
    niveis_cat = [nivel[i] for i in plot_df.index if plot_df.at[i, "Categoria"] == cat]
    n_niveis[cat] = max(niveis_cat) + 1 if niveis_cat else 1  # faixa vazia = 1 linha

# Quebra rótulos de faixa compridos em até 2 linhas, p/ a coluna da esquerda
# ficar mais estreita (em vez de reservar a largura do nome inteiro).
def quebra_rotulo(texto: str, largura: int = 15) -> str:
    linhas = textwrap.wrap(texto, width=largura, break_long_words=False)
    return "<br>".join(linhas) if linhas else texto

# Ordem das sub-linhas (de cima para baixo) e o texto que aparece à esquerda.
# O nome da faixa fica na sub-linha central; as demais ficam em branco.
ordem_lanes, rotulo_lane, faixa_intervalo = [], {}, {}
for cat in ordem_cats:
    inicio_idx = len(ordem_lanes)
    centro = (n_niveis[cat] - 1) // 2
    for n in range(n_niveis[cat]):
        chave = f"{cat}​{n}"  # ​ = caractere invisível só p/ deixar a chave única
        ordem_lanes.append(chave)
        rotulo_lane[chave] = quebra_rotulo(str(cat).upper()) if n == centro else ""
    faixa_intervalo[cat] = (inicio_idx, len(ordem_lanes) - 1)

plot_df["_lane"] = [f'{plot_df.at[i, "Categoria"]}​{nivel[i]}' for i in plot_df.index]

# Cor única POR LINHA (não por nome): duas ações com o mesmo nome — ex.: duas
# faixas "FRETE" — mantêm cada uma a sua cor, sem uma sobrescrever a outra.
plot_df["_cor_id"] = plot_df.index.astype(str)
mapa_cores = {}
for i, r in plot_df.iterrows():
    mapa_cores[str(i)] = r["Cor"] if isinstance(r["Cor"], str) and r["Cor"] else "#2E86DE"

# Descrição no tooltip: só entra (com quebra de linha) quando preenchida — assim
# não aparece "null" nas ações sem descrição.
_det = plot_df["Detalhes"].fillna("").astype(str).str.strip()
plot_df["_det_hover"] = _det.map(lambda d: f"<br>{d}" if d else "")

fig = px.timeline(
    plot_df,
    x_start="Início_plot",
    x_end="Fim_plot",
    y="_lane",
    color="_cor_id",
    color_discrete_map=mapa_cores,
    text="Rótulo",
    custom_data=["_det_hover", "Início_fmt", "Fim_fmt"],
    category_orders={"_lane": ordem_lanes},
)

fig.update_traces(
    textposition="auto",        # cabe na barra -> dentro; barra estreita -> ao lado
    insidetextanchor="middle",
    textangle=0,                # nunca gira o texto (sem escrita na vertical)
    cliponaxis=False,           # deixa o texto que sai da barra aparecer inteiro
    hovertemplate=(
        "<b>%{text}</b><br>"
        "Período: %{customdata[1]} → %{customdata[2]}"
        "%{customdata[0]}<extra></extra>"
    ),
)

# Cor do texto de cada barra conforme o fundo: preto em barras claras, branco em
# barras escuras — garante leitura nítida em qualquer cor da paleta.
for _tr in fig.data:
    _c = _tr.marker.color if isinstance(_tr.marker.color, str) else "#2E86DE"
    _tr.textfont = dict(color=_contraste_texto(_c), size=13)

# ===========================================================================
# Dois gráficos ALINHADOS: cabeçalho fixo (fig_head) + corpo rolável (fig)
# ===========================================================================
# O calendário é UM desenho só — não dá para "congelar" um pedaço dele. Então
# dividimos em DOIS gráficos Plotly que compartilham a MESMA margem esquerda/
# direita (l=120, r=10) e o MESMO intervalo de datas no eixo x, ficando
# encaixados coluna a coluna:
#   • fig_head -> faixa fininha do topo (Portfólio do Ciclo + SEMANAS + dias).
#                 Vai dentro de um container "grudado" no topo (position:sticky),
#                 então fica PARADO enquanto a página rola.
#   • fig      -> só as faixas (as barras), com a régua de datas ESCONDIDA.
# Assim dá para rolar até as faixas de baixo sem perder as datas de vista.

# --- Geometria compartilhada pelos dois gráficos ---------------------------
n_lanes = max(1, len(ordem_lanes))
mes_inicio = pd.Timestamp(ano, mes_num, 1)
mes_fim_excl = (
    pd.Timestamp(ano + 1, 1, 1) if mes_num == 12 else pd.Timestamp(ano, mes_num + 1, 1)
)
# Meio-dia de respiro em cada lado (os rótulos dos dias ficam centrados no marco
# da meia-noite; assim barras e faixa do ciclo casam com as bordas das células).
borda_esq = mes_inicio - meia
borda_dir = mes_fim_excl - meia
primeira_seg = mes_inicio - timedelta(days=int(mes_inicio.weekday()))

# Segmentos de semana (x0, x1, centro, número) — usados nos DOIS gráficos.
semanas_seg = []
_sem = 1
_d = primeira_seg
while _d < mes_fim_excl:
    _prox = _d + timedelta(days=7)
    _x0 = max(_d - meia, borda_esq)
    _x1 = min(_prox - meia, borda_dir)
    semanas_seg.append((_x0, _x1, _x0 + (_x1 - _x0) / 2, _sem))
    _d = _prox
    _sem += 1

# Régua de dias do mês (dia da semana em português; fim de semana em vermelho).
dias = pd.date_range(mes_inicio, mes_fim_excl - timedelta(days=1), freq="D")
PT_WD = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

def _rotulo_dia(x) -> str:
    """Rótulo 'DiaSemana<br>DD/MM'. Sáb (5) e Dom (6) saem em vermelho."""
    txt = f"{PT_WD[x.weekday()]}<br>{x:%d/%m}"
    if x.weekday() in (5, 6):
        return f"<span style='color:#E8463A'>{txt}</span>"
    return txt

# Ciclos visíveis no mês (faixa do Portfólio do Ciclo), em ordem de início.
ciclos = st.session_state.ciclos_df.dropna(subset=["Início", "Fim"])
ciclos_visiveis = []
for _, _c in ciclos.iterrows():
    _ci = pd.Timestamp(_c["Início"]) - meia
    _cf = pd.Timestamp(_c["Fim"]) + meia
    if _cf <= borda_esq or _ci >= borda_dir:
        continue  # ciclo fora do mês exibido
    ciclos_visiveis.append((_ci, _cf, str(_c["Ciclo"])))
ciclos_visiveis.sort(key=lambda t: t[0])

# --- Eixos e layout do CORPO (fig) -----------------------------------------
fig.update_yaxes(
    title="",
    showticklabels=False,  # rótulos das faixas viram anotações (alinhadas à direita)
    # IMPORTANTE: o px.timeline INVERTE a lista category_orders ao montar o eixo,
    # o que espelhava as barras em relação aos rótulos/linhas (que usam o índice
    # direto de ordem_lanes). Forçamos aqui o categoryarray na ordem correta, para
    # cada faixa cair exatamente na sua sub-linha (DESTAQUE no topo, ... CRM embaixo).
    categoryorder="array",
    categoryarray=ordem_lanes,
    autorange="reversed",  # 1ª faixa (DESTAQUE) no topo; última embaixo
    fixedrange=True,
)
fig.update_xaxes(
    title="",
    side="top",
    showgrid=False,          # sem linhas verticais dos dias
    showticklabels=False,    # a régua de datas fica SÓ no cabeçalho fixo
    range=[borda_esq, borda_dir],
    fixedrange=True,
)
# Altura do corpo: ~46px por sub-linha. Sem a margem de topo grande de antes
# (ciclo/semanas/dias foram para o cabeçalho); só um respiro mínimo em cima.
area_plot = 50 + 46 * n_lanes
fig.update_layout(
    showlegend=False,
    height=area_plot + 12,
    margin=dict(l=120, r=10, t=2, b=10),  # MESMO l/r do cabeçalho (colunas alinhadas)
    bargap=0,  # barras encostadas: sub-linhas da mesma faixa viram um bloco único
    dragmode=False,  # sem arrastar/zoom: o cursor não vira cruz no calendário
    paper_bgcolor="#FFFFFF",
    plot_bgcolor="#FFFFFF",
    # Tooltip (hover) bem legível: texto preto sobre fundo branco.
    hoverlabel=dict(
        bgcolor="#FFFFFF",
        bordercolor="#888888",
        font=dict(color="#000000", size=13),
        align="left",  # texto sempre alinhado à esquerda (não muda conforme a posição)
    ),
)

# Sombreado alternado das semanas — só na área das faixas (y de 0 a 1).
for _x0, _x1, _cx, _sn in semanas_seg:
    if _sn % 2 == 0:
        fig.add_shape(
            type="rect", xref="x", yref="paper",
            x0=_x0, x1=_x1, y0=0, y1=1,
            fillcolor="#000000", opacity=0.035, line_width=0, layer="below",
        )

# --- Linhas finas demarcando o topo e a base de cada faixa -----------------
# Cinza bem claro (~1px) fechando o bloco INTEIRO da faixa (todas as sub-linhas).
# No eixo categórico cada sub-linha ocupa um índice inteiro; as bordas entre
# faixas ficam nos meios (índice ± 0,5), dentro do "vão" entre as barras.
for cat in ordem_cats:
    ini_idx, fim_idx = faixa_intervalo[cat]
    for y in (ini_idx - 0.5, fim_idx + 0.5):
        fig.add_shape(
            type="line", xref="paper", yref="y",
            x0=-0.5, x1=1, y0=y, y1=y,  # x0 negativo: estende a linha até a coluna dos títulos
            line=dict(color="#E3E3E3", width=1),
            layer="above",
        )

# --- Linha BEM sutil entre as ações empilhadas (sub-linhas) da mesma faixa ---
# Só nos limites INTERNOS (os externos já têm a linha cinza acima). Fica sobre
# as barras, apenas na área do calendário (x de 0 a 1, sem a coluna de títulos).
for cat in ordem_cats:
    ini_idx, fim_idx = faixa_intervalo[cat]
    for k in range(ini_idx, fim_idx):        # limites internos entre sub-linhas
        fig.add_shape(
            type="line", xref="paper", yref="y",
            x0=0, x1=1, y0=k + 0.5, y1=k + 0.5,
            line=dict(color="rgba(0,0,0,0.22)", width=1),
            layer="above",
        )

# --- Rótulos das faixas (à esquerda) ---------------------------------------
# Como anotações alinhadas à DIREITA, para todos os títulos (faixas +
# "Portifólio do Ciclo") ficarem na mesma direção, encostados na coluna.
for cat in ordem_cats:
    ini_idx, fim_idx = faixa_intervalo[cat]
    rot = str(cat).upper()
    texto_rot = QUEBRA_MANUAL.get(rot, quebra_rotulo(rot))  # quebra manual tem prioridade
    fig.add_annotation(
        x=0, y=(ini_idx + fim_idx) / 2,
        xref="paper", yref="y",
        text=texto_rot,
        showarrow=False,
        font=dict(size=12, color="#444"),
        xanchor="right", yanchor="middle", align="right",
        xshift=-8,
    )

# ===========================================================================
# fig_head — CABEÇALHO FIXO (Portfólio do Ciclo + SEMANAS + régua de dias)
# ===========================================================================
# Figura própria só com o topo do calendário. Como é uma figura "vazia", o
# Plotly NÃO adivinha sozinho que o eixo X é de DATAS — então adicionamos uma
# linha invisível com as datas (âncora) e forçamos type="date". Assim a régua
# de dias e a faixa do ciclo se espalham pela largura, alinhadas ao corpo.
# Layout de cima p/ baixo: faixa do ciclo -> "SEMANA N" -> régua de dias. Os
# dias ficam na margem DE BAIXO (side="bottom"), encostados no corpo; ciclo e
# semanas ficam DENTRO da área do gráfico (y de 0 a 1), sem coords extremas.
fig_head = go.Figure()
# Âncora invisível: dá ao eixo X os limites de data (sem desenhar nada visível).
fig_head.add_trace(go.Scatter(
    x=[borda_esq, borda_dir], y=[0, 0], mode="markers",
    marker=dict(opacity=0), hoverinfo="skip", showlegend=False,
))
fig_head.update_layout(
    height=104,
    margin=dict(l=120, r=10, t=6, b=40),  # MESMO l/r do corpo; b p/ os rótulos dos dias
    showlegend=False,
    dragmode=False,
    paper_bgcolor="#FFFFFF",
    plot_bgcolor="#FFFFFF",
    xaxis=dict(
        type="date",
        range=[borda_esq, borda_dir], fixedrange=True, side="bottom",
        showgrid=False, showline=False, ticks="",
        tickmode="array", tickvals=list(dias),
        ticktext=[_rotulo_dia(x) for x in dias],
        tickfont=dict(color="#333333", size=12),
    ),
    yaxis=dict(range=[0, 1], visible=False, fixedrange=True),
)

# Sombreado alternado das semanas (cobre a área do cabeçalho, descendo um pouco
# sobre os rótulos dos dias) + rótulo "SEMANA N" centralizado.
for _x0, _x1, _cx, _sn in semanas_seg:
    if _sn % 2 == 0:
        fig_head.add_shape(
            type="rect", xref="x", yref="paper",
            x0=_x0, x1=_x1, y0=-0.8, y1=1.0,
            fillcolor="#000000", opacity=0.035, line_width=0, layer="below",
        )
    fig_head.add_annotation(
        x=_cx, y=0.26, yref="paper", text=f"SEMANA {_sn}", showarrow=False,
        font=dict(size=12, color="#333333"), xanchor="center", yanchor="middle",
    )

# Faixa "Portfólio do Ciclo" (amarela/azul alternadas), no topo da área.
for idx, (_ci, _cf, _texto) in enumerate(ciclos_visiveis):
    cor = CORES_CICLO[idx % len(CORES_CICLO)]  # alterna amarelo / azul / ...
    _x0 = max(_ci, borda_esq)
    _x1 = min(_cf, borda_dir)
    fig_head.add_shape(
        type="rect", xref="x", yref="paper",
        x0=_x0, x1=_x1, y0=0.55, y1=1.0,
        fillcolor=cor["fill"], line=dict(color=cor["borda"], width=1),
        layer="above",
    )
    fig_head.add_annotation(
        x=_x0 + (_x1 - _x0) / 2, y=0.775, xref="x", yref="paper",
        text=f"<b>{_texto}</b>", showarrow=False,
        font=dict(size=12, color=_contraste_texto(cor["fill"])),
        xanchor="center", yanchor="middle",
    )

# Rótulo fixo "PORTIFÓLIO DO CICLO" à esquerda da faixa amarela.
fig_head.add_annotation(
    x=0, y=0.775, xref="paper", yref="paper",
    text="PORTIFÓLIO<br>DO CICLO", showarrow=False,
    font=dict(size=12, color="#333"),
    xanchor="right", yanchor="middle", align="right", xshift=-8,
)

# ===========================================================================
# fig_destaque — a faixa "DESTAQUE DA COMUNICAÇÃO", GRUDADA no cabeçalho
# ===========================================================================
# Terceiro gráfico, com a MESMA margem (l=120, r=10) e o MESMO intervalo de datas
# dos outros dois — então as colunas casam. Vai na zona fixa (não rola). Como esta
# faixa NUNCA empilha (sempre uma comunicação após a outra), é uma sub-linha só.
topo_df["_lane"] = FAIXA_FIXA_TOPO  # uma única sub-linha
topo_df["_cor_id"] = topo_df.index.astype(str)
mapa_cores_topo = {
    str(i): (r["Cor"] if isinstance(r["Cor"], str) and r["Cor"] else "#2E86DE")
    for i, r in topo_df.iterrows()
}
_det_topo = topo_df["Detalhes"].fillna("").astype(str).str.strip()
topo_df["_det_hover"] = _det_topo.map(lambda d: f"<br>{d}" if d else "")

fig_destaque = px.timeline(
    topo_df,
    x_start="Início_plot", x_end="Fim_plot", y="_lane",
    color="_cor_id", color_discrete_map=mapa_cores_topo,
    text="Rótulo",
    custom_data=["_det_hover", "Início_fmt", "Fim_fmt"],
)
fig_destaque.update_traces(
    textposition="auto", insidetextanchor="middle", textangle=0, cliponaxis=False,
    hovertemplate=(
        "<b>%{text}</b><br>"
        "Período: %{customdata[1]} → %{customdata[2]}"
        "%{customdata[0]}<extra></extra>"
    ),
)
for _tr in fig_destaque.data:
    _c = _tr.marker.color if isinstance(_tr.marker.color, str) else "#2E86DE"
    _tr.textfont = dict(color=_contraste_texto(_c), size=13)

fig_destaque.update_yaxes(title="", showticklabels=False, fixedrange=True)
fig_destaque.update_xaxes(
    title="", showgrid=False, showticklabels=False,
    range=[borda_esq, borda_dir], fixedrange=True,
)
fig_destaque.update_layout(
    showlegend=False,
    height=58,                                   # ~46px de barra (igual às do corpo)
    margin=dict(l=120, r=10, t=6, b=6),          # MESMO l/r -> colunas alinhadas
    bargap=0,
    dragmode=False,
    paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
    hoverlabel=dict(bgcolor="#FFFFFF", bordercolor="#888888",
                    font=dict(color="#000000", size=13), align="left"),
)
# Sombreado alternado das semanas (igual ao corpo e ao cabeçalho).
for _x0, _x1, _cx, _sn in semanas_seg:
    if _sn % 2 == 0:
        fig_destaque.add_shape(
            type="rect", xref="x", yref="paper", x0=_x0, x1=_x1, y0=0, y1=1,
            fillcolor="#000000", opacity=0.035, line_width=0, layer="below",
        )
# Linhas finas fechando o topo e a base da faixa (mesmo cinza das faixas do corpo).
for _yb in (0, 1):
    fig_destaque.add_shape(
        type="line", xref="paper", yref="paper", x0=-0.5, x1=1, y0=_yb, y1=_yb,
        line=dict(color="#E3E3E3", width=1), layer="above",
    )
# Rótulo "DESTAQUE DA COMUNICAÇÃO" à esquerda (mesma direção dos demais títulos).
fig_destaque.add_annotation(
    x=0, y=0.5, xref="paper", yref="paper",
    text=quebra_rotulo(FAIXA_FIXA_TOPO.upper()), showarrow=False,
    font=dict(size=12, color="#444"),
    xanchor="right", yanchor="middle", align="right", xshift=-8,
)

# --- Render: DOIS quadros — cabeçalho fixo em cima + faixas roláveis embaixo -
# O position:sticky não cola de forma confiável no Streamlit (nem solto na
# página, nem dentro de um quadro). Solução determinística: separar em DOIS
# quadros irmãos. O cabeçalho (datas) fica num quadro PRÓPRIO que NÃO rola; as
# faixas ficam num quadro logo abaixo com rolagem própria (overflow-y: auto).
# Como o cabeçalho é um elemento à parte, ACIMA do quadro que rola, ele
# simplesmente não se mexe quando você rola as faixas. O `scrollbar-gutter:
# stable` reserva o MESMO espaço de barra de rolagem nos dois, para as colunas
# de datas ficarem alinhadas com as barras das faixas.
# O calendário PREENCHE a largura no computador, mas NUNCA fica mais estreito que
# um mínimo legível. O truque que funciona no Streamlit: o MÍNIMO vai no QUADRO
# (não dentro do gráfico), e cabeçalho + faixas ficam dentro de UM ÚNICO envelope
# que rola na horizontal — assim os dois rolam JUNTOS e continuam alinhados.
# LARGURA_MIN = 120+10 de margem + ~36px por dia. Ajuste o "36" (px por dia).
LARGURA_MIN = 130 + 36 * len(dias)

st.markdown(
    f"""
    <style>
      /* Envelope único que rola na horizontal (arrasta pro lado no celular). */
      .st-key-cal_scroll {{ overflow-x: auto; }}
      /* MÍNIMO legível no quadro: no PC ele cresce e preenche; no celular fica no
         mínimo e o envelope acima rola. Cabeçalho e faixas têm o mesmo mínimo,
         então ficam sempre alinhados. */
      .st-key-cal_head_box, .st-key-cal_box {{
        min-width: {LARGURA_MIN}px;
        scrollbar-gutter: stable;               /* mesmo recuo de barra nos dois */
        border-left: 1px solid #ECECEC;
        border-right: 1px solid #ECECEC;
      }}
      /* Cabeçalho (datas): min-height evita que ele encolha em telas baixas e
         corte a linha "SEMANA N" do topo. */
      .st-key-cal_head_box {{
        overflow-y: auto;
        min-height: 112px;
        border-top: 1px solid #ECECEC;
        border-radius: 6px 6px 0 0;
        margin-bottom: -1rem;
        gap: 0 !important;   /* cola a faixa fixa logo abaixo das datas, sem vão */
      }}
      /* Faixas: rolam na vertical aqui dentro. */
      .st-key-cal_box {{
        max-height: 68vh;
        overflow-y: auto;
        border-bottom: 1px solid #ECECEC;
        border-radius: 0 0 6px 6px;
      }}
    </style>
    """,
    unsafe_allow_html=True,
)

_cfg_cal = {
    "displayModeBar": False,      # esconde a barra de ferramentas (canto sup. dir.)
    "scrollZoom": False,          # sem zoom pela rolagem do mouse
    "doubleClick": False,         # sem reset/zoom no duplo-clique
    "showAxisDragHandles": False, # sem alças de arraste nos eixos
    "displaylogo": False,
}
# Um ÚNICO envelope rolável na horizontal, com o cabeçalho em cima e as faixas
# embaixo — os dois com o MESMO piso de largura, então sempre alinhados ao rolar.
with st.container(key="cal_scroll"):
    with st.container(key="cal_head_box"):
        st.plotly_chart(
            fig_head, use_container_width=True, config=_cfg_cal,
            key="grafico_cabecalho",
        )
        # A faixa fixa "DESTAQUE DA COMUNICAÇÃO", logo abaixo das datas (não rola).
        st.plotly_chart(
            fig_destaque, use_container_width=True, config=_cfg_cal,
            key="grafico_destaque",
        )
    with st.container(key="cal_box"):
        st.plotly_chart(
            fig, use_container_width=True, config=_cfg_cal, key="grafico_corpo",
        )


# ----------------------------------------------------------------------------
# Tabela: editar / excluir
# ----------------------------------------------------------------------------
def _swatch(cor) -> str:
    """Quadradinho SVG (data-URI) preenchido com a cor do hex — prévia visual."""
    c = cor.strip() if isinstance(cor, str) else ""
    fill = c.replace("#", "%23") if c else "none"  # '#' precisa ser %23 na URI
    return (
        "data:image/svg+xml;utf8,"
        "<svg xmlns='http://www.w3.org/2000/svg' width='36' height='20'>"
        f"<rect x='1' y='1' width='34' height='18' rx='4' fill='{fill}' "
        "stroke='%23BBBBBB' stroke-width='1'/></svg>"
    )


# Leitor não edita: a tabela vira só leitura e sobra só o "Baixar Excel".
_titulo_painel = "✏️ Editar ou excluir ações" if PODE_EDITAR else "📋 Ações do mês / baixar Excel"
with st.expander(_titulo_painel, expanded=False):
    if PODE_EDITAR:
        st.caption(
            "Mostra as ações do **mês selecionado** no topo. Edite qualquer célula direto na "
            "tabela. Para **excluir**, marque a caixa **🗑️ Excluir** da(s) linha(s) e clique "
            "em **Salvar alterações**."
        )
    else:
        st.caption(
            "Mostra as ações do **mês selecionado**. Você pode **baixar o Excel** aqui embaixo. "
            "(Edição disponível apenas para administradores/editores.)"
        )

    # Mostra só as ações do mês selecionado (igual ao calendário). As ações dos
    # outros meses ficam em 'outros' e são regravadas ao salvar (não se perdem).
    df_full = st.session_state.df
    if df_full.empty:
        editor_df = df_full.copy()
        outros = df_full.copy()
    else:
        overlap = (df_full["Início"] <= sel_fim) & (df_full["Fim"] >= sel_inicio)
        sem_data = df_full["Início"].isna() | df_full["Fim"].isna()
        mask_mes = overlap | sem_data
        editor_df = df_full[mask_mes].copy()
        outros = df_full[~mask_mes].copy()

    # Coluna só de exibição com o quadradinho da cor (descartada ao salvar).
    editor_df["🎨"] = editor_df["Cor"].map(_swatch)
    # Checkbox visível para marcar linhas a excluir (aplicado ao salvar).
    editor_df["Excluir"] = False

    # Garante que as colunas de texto sejam string. Quando o mês não tem nenhum
    # valor preenchido (ex.: "Detalhes" vazio em todas as linhas), o pandas infere
    # o tipo como float e o st.data_editor gera StreamlitAPIException ao casar com
    # a TextColumn. Preenche NaN com "" e força str para manter a compatibilidade.
    for _col in ("Ação", "Categoria", "Cor", "Detalhes"):
        if _col in editor_df.columns:
            editor_df[_col] = editor_df[_col].fillna("").astype(str)

    # Numera a tabela DO INÍCIO em cada mês. Sem isto, o número da lateral herda a
    # posição na lista geral (todos os meses juntos) e um mês começa no 13, outro
    # no 20, etc. Reinicia a contagem começando no 1 (1, 2, 3...) em TODO mês.
    editor_df = editor_df.reset_index(drop=True)
    editor_df.index = editor_df.index + 1

    # Opções do dropdown de Categoria: faixas padrão + as já usadas nas ações
    # (inclui todo valor presente para o SelectboxColumn não gerar erro).
    _cat_opcoes = sorted(set(CATEGORIAS_PADRAO) | set(editor_df["Categoria"]))

    editado = st.data_editor(
        editor_df,
        num_rows="fixed",  # sem a seleção embutida do Streamlit (usamos a coluna "Excluir")
        disabled=not PODE_EDITAR,  # leitor: tabela só leitura (não altera nada)
        use_container_width=True,
        column_order=["Excluir", "Ação", "Categoria", "Início", "Fim", "🎨", "Cor", "Detalhes"],
        column_config={
            "Excluir": st.column_config.CheckboxColumn(
                "🗑️ Excluir",
                help="Marque e clique em Salvar alterações para apagar a linha.",
                default=False,
            ),
            "Ação": st.column_config.TextColumn("Ação", required=True),
            "Categoria": st.column_config.SelectboxColumn(
                "Categoria",
                options=_cat_opcoes,
                help="Selecione a faixa (lado esquerdo do calendário).",
            ),
            "Início": st.column_config.DateColumn("Início", format="DD/MM/YYYY"),
            "Fim": st.column_config.DateColumn("Fim", format="DD/MM/YYYY"),
            "🎨": st.column_config.ImageColumn("Cor", help="Prévia da cor do hex ao lado"),
            "Cor": st.column_config.TextColumn("Cor (hex)", help="Ex: #E8463A"),
            "Detalhes": st.column_config.TextColumn("Detalhes", width="large"),
        },
        key="editor",
    )

    col_a, col_b, col_c = st.columns([1, 1, 2])
    with col_a:
        if PODE_EDITAR and st.button("💾 Salvar alterações", type="primary"):
            # Remove as linhas marcadas em "Excluir", descarta as colunas auxiliares
            # (🎨 e Excluir) e junta com as ações dos outros meses preservadas.
            salvas = editado.drop(columns=["🎨"], errors="ignore")
            if "Excluir" in salvas.columns:
                manter = ~salvas["Excluir"].fillna(False).astype(bool)
                salvas = salvas[manter].drop(columns=["Excluir"])
            st.session_state.df = pd.concat([outros, salvas], ignore_index=True)[COLUNAS]
            salvar(st.session_state.df)
            st.success("Alterações salvas!")
            st.rerun()

    with col_b:
        # Gera o arquivo Excel (.xlsx) em memória, já formatado, para download.
        buffer = io.BytesIO()
        # "Cor" (hex) não é relevante no relatório — fica de fora do Excel.
        df_export = st.session_state.df.drop(columns=["Cor"], errors="ignore").copy()
        # Datas como datetime (para o Excel reconhecer e formatar dd/mm/aaaa).
        for _dcol in ("Início", "Fim"):
            if _dcol in df_export.columns:
                df_export[_dcol] = pd.to_datetime(df_export[_dcol], errors="coerce")

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Ações")
            ws = writer.sheets["Ações"]

            # --- Estilos ---
            verde = PatternFill("solid", fgColor="A9D18E")    # cabeçalho
            fonte_cab = Font(bold=True, color="1F3B08")
            centro = Alignment(horizontal="center", vertical="center", wrap_text=False)
            esq = Alignment(horizontal="left", vertical="center", wrap_text=False)
            _fina = Side(style="thin", color="BFBFBF")
            borda = Border(left=_fina, right=_fina, top=_fina, bottom=_fina)

            cols = list(df_export.columns)
            idx = {nome: cols.index(nome) + 1 for nome in cols}
            centralizar = {"Categoria", "Início", "Fim", "Cor"}
            n_lin, n_col = ws.max_row, ws.max_column

            # Cabeçalho (linha 1): verde, negrito, centralizado.
            for c in range(1, n_col + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = verde
                cell.font = fonte_cab
                cell.alignment = centro
                cell.border = borda

            # Corpo: bordas, alinhamento e datas dd/mm/aaaa.
            for r in range(2, n_lin + 1):
                for c in range(1, n_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = borda
                    cell.alignment = centro if cols[c - 1] in centralizar else esq
                for _dcol in ("Início", "Fim"):
                    if _dcol in idx:
                        ws.cell(row=r, column=idx[_dcol]).number_format = "DD/MM/YYYY"

            # Largura automática das colunas (datas com largura fixa).
            for c in range(1, n_col + 1):
                letra = get_column_letter(c)
                if cols[c - 1] in ("Início", "Fim"):
                    ws.column_dimensions[letra].width = 12
                    continue
                maior = max(
                    (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, n_lin + 1)),
                    default=10,
                )
                ws.column_dimensions[letra].width = min(max(maior + 2, 10), 55)

            ws.freeze_panes = "A2"  # cabeçalho fixo ao rolar

        st.download_button(
            "⬇️ Baixar Excel",
            data=buffer.getvalue(),
            file_name="acoes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with col_c:
        # Limpar TODO o calendário — no canto direito, discreto. Uso raro (recomeçar
        # do zero), mas disponível. Só admin/editor. Pede confirmação antes de apagar.
        if PODE_EDITAR and not st.session_state.get("confirmar_limpeza"):
            if st.button("🗑️ Limpar calendário (apagar tudo)"):
                st.session_state.confirmar_limpeza = True
                st.rerun()

    # Confirmação da limpeza — aparece só após clicar, dentro deste painel.
    if PODE_EDITAR and st.session_state.get("confirmar_limpeza"):
        st.warning(
            "Tem certeza? Isso apaga **TODAS** as ações do calendário no banco. "
            "Essa ação não tem como desfazer aqui pelo app."
        )
        c1, c2, _ = st.columns([1, 1, 4])
        with c1:
            if st.button("✅ Sim, apagar tudo", type="primary"):
                st.session_state.df = pd.DataFrame(columns=COLUNAS)
                salvar(st.session_state.df)  # grava calendário vazio no Supabase
                st.session_state.confirmar_limpeza = False
                st.session_state.flash = "Calendário limpo!"
                st.rerun()
        with c2:
            if st.button("Cancelar"):
                st.session_state.confirmar_limpeza = False
                st.rerun()
